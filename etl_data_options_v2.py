#!/usr/bin/env python3
"""
ETL v2: Parse data-options-cleaned.xlsx → normalized records
Output: area, store_name, display_type, gender, kategori, capacity
"""

import openpyxl
import csv, sys

EXCEL_PATH = '/Users/database-zuma/.openclaw/workspace/data-options-cleaned.xlsx'
OUTPUT_CSV  = '/Users/database-zuma/.openclaw/workspace/data-options-upload.csv'

SKIP_SHEETS   = ['Rules']
SKIP_LABELS   = {'Grand Total', 'Nama Toko', None}

def clean_str(v):
    if v is None: return None
    s = str(v).strip()
    return s if s else None

def is_store_name(v):
    """Return True if value looks like a store name (not a number, not 'Gender'/'Kategori')"""
    if v is None: return False
    s = clean_str(v)
    if not s: return False
    upper = s.upper()
    if upper in ('GENDER', 'KATEGORI', 'GRAND TOTAL', 'NAMA TOKO'): return False
    try:
        float(s); return False  # it's numeric
    except (ValueError, TypeError): pass
    return True

def clean_num(v):
    if v is None: return None
    try:
        f = float(v)
        return int(f) if f == int(f) else round(f, 4)
    except: return None


def find_row_by_label(rows, label):
    """Return (index, row) where rows[idx][0] == label, or (None, None)"""
    for i, row in enumerate(rows):
        if clean_str(row[0]) == label:
            return i, row
    return None, None


def parse_standard_sheet(ws, area):
    """
    Handles: Bali 1, Bali 2, Bali 3, Jakarta, Jatim, Sumatera, Lombok
    Store names live in either:
      - 'Grand Total' row (at cols 1,4,7,...) → Type B
      - 'Nama Toko' row  (at cols 1,4,7,...)  → Type A (Bali 1, Lombok)
    Data rows: [label, gender_s1, kat_s1, val_s1, gender_s2, kat_s2, val_s2, ...]
    """
    rows = list(ws.iter_rows(values_only=True))
    gt_idx, gt_row = find_row_by_label(rows, 'Grand Total')
    nt_idx, nt_row = find_row_by_label(rows, 'Nama Toko')

    # Determine store-name row
    stores = {}  # group_idx -> store_name
    store_src = 'Grand Total'

    if gt_row is not None:
        for col in range(1, len(gt_row), 3):
            val = gt_row[col]
            if is_store_name(val):
                stores[(col - 1) // 3] = clean_str(val)

    if not stores and nt_row is not None:
        store_src = 'Nama Toko'
        for col in range(1, len(nt_row), 3):
            val = nt_row[col]
            if is_store_name(val):
                stores[(col - 1) // 3] = clean_str(val)

    if not stores:
        # Try Nama Toko anyway (Lombok-style: stores at 1,4 in nt_row but different format)
        if nt_row is not None:
            store_src = 'Nama Toko'
            for col in range(1, len(nt_row)):
                val = nt_row[col]
                if is_store_name(val):
                    group = (col - 1) // 3
                    stores[group] = clean_str(val)

    print(f"  [{area}] store source='{store_src}', stores={list(stores.values())}")

    # Data starts after the last header row
    data_start = max(
        gt_idx if gt_idx is not None else 0,
        nt_idx if nt_idx is not None else 0
    ) + 1

    records = []
    for row in rows[data_start:]:
        label = clean_str(row[0])
        if not label or label in SKIP_LABELS:
            continue

        for group_idx, store_name in stores.items():
            base = 1 + group_idx * 3
            if base + 2 >= len(row):
                continue
            gender   = clean_str(row[base])
            kategori = clean_str(row[base + 1])
            capacity = clean_num(row[base + 2])

            if capacity is None and gender is None and kategori is None:
                continue

            records.append({
                'area': area, 'store_name': store_name,
                'display_type': label, 'gender': gender,
                'kategori': kategori, 'capacity': capacity,
            })
    return records


def parse_sulawesi(ws, area):
    """
    Sulawesi: different layout
    Row 'Nama Toko': [label, 'Gender', 'Kategori', store1, store2, ...]
    Data rows: [label, None, None, val1, val2, ...]
    No gender/kategori per store.
    """
    rows = list(ws.iter_rows(values_only=True))
    nt_idx, nt_row = find_row_by_label(rows, 'Nama Toko')

    if nt_row is None:
        print(f"  [{area}] WARNING: no Nama Toko row found", file=sys.stderr)
        return []

    # Stores at cols 3, 4, 5, ... (after Gender at 1, Kategori at 2)
    stores = {}  # col_idx -> store_name
    for col in range(3, len(nt_row)):
        val = nt_row[col]
        if is_store_name(val):
            stores[col] = clean_str(val)

    print(f"  [{area}] stores (flat cols)={list(stores.items())}")

    records = []
    for row in rows[nt_idx + 1:]:
        label = clean_str(row[0])
        if not label or label in SKIP_LABELS:
            continue
        for col, store_name in stores.items():
            if col >= len(row): continue
            capacity = clean_num(row[col])
            if capacity is None: continue
            records.append({
                'area': area, 'store_name': store_name,
                'display_type': label, 'gender': None,
                'kategori': None, 'capacity': capacity,
            })
    return records


def parse_consignment(ws, area):
    """
    Consignment: one column per store, no gender/kategori
    Row 'Nama Toko': [label, store1, store2, ...]
    Data rows: [label, val1, val2, ...]
    """
    rows = list(ws.iter_rows(values_only=True))
    nt_idx, nt_row = find_row_by_label(rows, 'Nama Toko')

    stores = {}  # col_idx -> store_name
    for col in range(1, len(nt_row)):
        val = nt_row[col]
        if is_store_name(val):
            stores[col] = clean_str(val)

    print(f"  [{area}] {len(stores)} stores")

    records = []
    for row in rows[nt_idx + 1:]:
        label = clean_str(row[0])
        if not label or label in SKIP_LABELS:
            continue
        for col, store_name in stores.items():
            if col >= len(row): continue
            capacity = clean_num(row[col])
            if capacity is None: continue
            records.append({
                'area': area, 'store_name': store_name,
                'display_type': label, 'gender': None,
                'kategori': None, 'capacity': capacity,
            })
    return records


def main():
    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
    all_records = []

    for sheet_name in wb.sheetnames:
        if sheet_name in SKIP_SHEETS:
            continue

        ws = wb[sheet_name]
        print(f"\nParsing: {sheet_name}")

        if sheet_name == 'Sulawesi':
            records = parse_sulawesi(ws, sheet_name)
        elif sheet_name == 'Consignment':
            records = parse_consignment(ws, sheet_name)
        else:
            records = parse_standard_sheet(ws, sheet_name)

        print(f"  -> {len(records)} records")
        all_records.extend(records)

    print(f"\n{'='*50}")
    print(f"Total records: {len(all_records)}")

    # Write CSV
    fields = ['area', 'store_name', 'display_type', 'gender', 'kategori', 'capacity']
    with open(OUTPUT_CSV, 'w', newline='', encoding='utf-8') as f:
        writer = csv.DictWriter(f, fieldnames=fields)
        writer.writeheader()
        writer.writerows(all_records)

    print(f"CSV: {OUTPUT_CSV}")

    from collections import Counter
    area_counts = Counter(r['area'] for r in all_records)
    store_counts = Counter(r['store_name'] for r in all_records)
    print(f"\nRecords per area:")
    for k, v in sorted(area_counts.items()):
        print(f"  {k}: {v}")
    print(f"\nUnique stores: {len(store_counts)}")
    for k, v in sorted(store_counts.items()):
        print(f"  {k}: {v} rows")


if __name__ == '__main__':
    main()

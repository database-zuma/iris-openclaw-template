#!/usr/bin/env python3
"""
Clean "Data Options" Excel file from Wayan.
- Keep only "new" sheet variants, drop old ones + Sheet6
- Fix typos, trailing spaces, capitalization issues, store names
"""

import openpyxl
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
import re

SRC = '/Users/database-zuma/.openclaw/media/inbound/1227256a-9c3a-4639-b465-a564c51628d8.xlsx'
DST = '/Users/database-zuma/.openclaw/workspace/data-options-cleaned.xlsx'

# ── Sheet selection: old_name → new_name (keep only these) ──────────────────
KEEP_SHEETS = {
    'Rules':        'Rules',
    'Bali 1 New':   'Bali 1',
    'Bali 2 New ':  'Bali 2',
    'Bali 3 New ':  'Bali 3',
    'Lombok':       'Lombok',
    'Jakarta NEW':  'Jakarta',
    'Jatim New':    'Jatim',
    'Sumatera':     'Sumatera',
    'Sulawesi':     'Sulawesi',
    'Consignment':  'Consignment',
}

DROPPED_SHEETS = ['Sheet6', 'Bali 1', 'Bali 2', 'Bali 3', 'Jakarta', 'Jatim']

# ── Store name corrections (exact match after strip) ────────────────────────
STORE_NAME_FIXES = {
    'ZUMA BAJERA':       'ZUMA BAJRA',          # typo extra E
    'ZUMA LEVEL21':      'ZUMA LEVEL 21',        # missing space
    'ZUMA UBUDMF':       'ZUMA MF UBUD',         # word order wrong
    'ZUMA UBUDMF ':      'ZUMA MF UBUD',         # trailing space variant
    'ZUMA LIPPOMALL BALI': 'ZUMA LIPPO BALI',    # MALL not part of name
    'PELIATAN':          'ZUMA PELIATAN',         # missing ZUMA prefix
    'ZUMA Nagoya Hill':  'ZUMA NAGOYA HILL',      # inconsistent caps
    'Zuma Pepito kutuh': 'Zuma Pepito Kutuh',     # lowercase K
    'Aeon Cakung ':      'Aeon Cakung',           # trailing space in store
    'Dalung':            'DALUNG',               # inconsistent caps (row-0 abbrev)
}

# ── Keyword/column header normalisation ─────────────────────────────────────
# After strip, these exact strings → normalised form
HEADER_FIXES = {
    'kategorie':  'Kategori',
    'Kategorie':  'Kategori',
    'kategori':   'Kategori',
    'KATEGORI':   'Kategori',
    'gender':     'Gender',
    'GENDER':     'Gender',
    'GENDER ':    'Gender',   # space variant
    ' Nama Toko': 'Nama Toko',
}

# ── Typo word-level fixes (applied to every string cell) ───────────────────
# Each entry: (regex_pattern, replacement_for_upper, replacement_for_title, replacement_for_lower)
# We'll detect case from the matched token.

def case_preserving_sub(pattern, correct_word, text):
    """Replace `pattern` (case-insensitive) with correct_word, preserving UPPER / Title / lower style."""
    def _replace(m):
        w = m.group(0)
        if w.isupper():
            return correct_word.upper()
        elif w.istitle() or (w[0].isupper() and not w.isupper()):
            return correct_word.capitalize()
        else:
            return correct_word.lower()
    return re.sub(pattern, _replace, text, flags=re.IGNORECASE)

def fix_typos(s):
    """Apply all typo corrections to a string value."""
    # 1. Fix 'japit' → 'jepit' (word boundary aware)
    s = case_preserving_sub(r'\bjapit\b', 'Jepit', s)

    # 2. Fix 'fashioj' → 'fashion'
    s = case_preserving_sub(r'\bfashioj\b', 'Fashion', s)

    # 3. Fix 'fashioh' → 'fashion'
    s = case_preserving_sub(r'\bfashioh\b', 'Fashion', s)

    # 4. Fix 'FASION' / 'fasion' → 'fashion' (missing H)
    s = case_preserving_sub(r'\bfasion\b', 'Fashion', s)

    # 5. Fix 'Girs' → 'Girls'
    s = re.sub(r'\bGirs\b', 'Girls', s)

    # 6. Fix 'clasic' → 'classic'
    s = re.sub(r'\bclasic\b', 'classic', s, flags=re.IGNORECASE)

    # 7. Fix '80B0X' (zero substituted for letter O)
    s = s.replace('80B0X', '80BOX')

    return s

def clean_cell_value(val):
    """Apply all cleaning to a single cell value. Returns (cleaned_val, was_changed, description)."""
    if not isinstance(val, str):
        return val, False, None

    original = val
    changes = []

    # Step 1: Strip leading/trailing whitespace
    stripped = val.strip()
    if stripped != val:
        changes.append(f'strip whitespace: {repr(val)} → {repr(stripped)}')
        val = stripped

    # Step 2: Store name fixes (exact match after strip)
    if val in STORE_NAME_FIXES:
        fixed = STORE_NAME_FIXES[val]
        changes.append(f'store name: {repr(val)} → {repr(fixed)}')
        val = fixed

    # Step 3: Header keyword normalisation (exact match)
    if val in HEADER_FIXES:
        fixed = HEADER_FIXES[val]
        changes.append(f'header: {repr(val)} → {repr(fixed)}')
        val = fixed

    # Step 4: Typo fixes (substring/word level)
    fixed = fix_typos(val)
    if fixed != val:
        changes.append(f'typo: {repr(val)} → {repr(fixed)}')
        val = fixed

    was_changed = val != original
    return val, was_changed, '; '.join(changes) if changes else None


# ────────────────────────────────────────────────────────────────────────────
# Main processing
# ────────────────────────────────────────────────────────────────────────────

print("Loading source file (data_only=True)…")
wb_src = openpyxl.load_workbook(SRC, data_only=True)

dest_wb = Workbook()
dest_wb.remove(dest_wb.active)   # remove default blank sheet

change_log = []
sheet_summary = []

for old_name, new_name in KEEP_SHEETS.items():
    if old_name not in wb_src.sheetnames:
        print(f"  WARNING: sheet {old_name!r} not found, skipping")
        continue

    ws_src = wb_src[old_name]
    ws_dst = dest_wb.create_sheet(title=new_name)

    data_row_count = 0
    cell_changes = 0

    for row in ws_src.iter_rows():
        for cell in row:
            val = cell.value

            # ── skip formula-error strings that data_only cached ──────────
            if isinstance(val, str) and val.strip() in ('#REF!', '#DIV/0!', '#VALUE!', '#NAME?', '#N/A'):
                val = None   # clear error artifacts

            cleaned_val, changed, desc = clean_cell_value(val)

            # Write value to destination
            dst_cell = ws_dst.cell(row=cell.row, column=cell.column, value=cleaned_val)

            # Copy basic style info (font, fill, alignment) if possible
            if cell.has_style:
                try:
                    from copy import copy
                    dst_cell.font      = copy(cell.font)
                    dst_cell.fill      = copy(cell.fill)
                    dst_cell.border    = copy(cell.border)
                    dst_cell.alignment = copy(cell.alignment)
                    dst_cell.number_format = cell.number_format
                except Exception:
                    pass

            if changed:
                cell_changes += 1
                change_log.append({
                    'sheet': new_name,
                    'cell': f'{get_column_letter(cell.column)}{cell.row}',
                    'detail': desc
                })

        if any(c.value is not None for c in row):
            data_row_count += 1

    # Copy column widths
    for col_dim in ws_src.column_dimensions.values():
        ws_dst.column_dimensions[col_dim.index].width = col_dim.width

    sheet_summary.append({
        'sheet': new_name,
        'original': old_name,
        'rows': data_row_count,
        'cells_fixed': cell_changes,
    })
    print(f"  ✓ {old_name!r} → {new_name!r}  ({data_row_count} rows, {cell_changes} cells fixed)")

print(f"\nSaving to {DST}…")
dest_wb.save(DST)
print("Done.\n")

# ────────────────────────────────────────────────────────────────────────────
# Print report
# ────────────────────────────────────────────────────────────────────────────
print("=" * 70)
print("CLEANING REPORT")
print("=" * 70)

print("\n📋 SHEETS SELECTED / DROPPED")
print(f"  Kept ({len(KEEP_SHEETS)}):   " + ", ".join(KEEP_SHEETS.values()))
print(f"  Dropped ({len(DROPPED_SHEETS)}): " + ", ".join(DROPPED_SHEETS))

print("\n📊 FINAL SHEET STRUCTURE")
print(f"  {'Sheet':<12} {'Original Name':<18} {'Data Rows':>10} {'Cells Fixed':>12}")
print(f"  {'-'*12} {'-'*18} {'-'*10} {'-'*12}")
for s in sheet_summary:
    renamed = ' (renamed)' if s['sheet'] != s['original'].strip() else ''
    print(f"  {s['sheet']:<12} {s['original']:<18} {s['rows']:>10} {s['cells_fixed']:>12}")

print(f"\n🔧 CHANGES MADE (first 60)")
for i, c in enumerate(change_log[:60]):
    print(f"  [{c['sheet']} {c['cell']}] {c['detail']}")
if len(change_log) > 60:
    print(f"  … and {len(change_log)-60} more")

print(f"\n  Total cells changed: {len(change_log)}")

print("\n📝 RENAME MAP (new → old)")
for old, new in KEEP_SHEETS.items():
    if old.strip() != new:
        print(f"  {old!r:25} → {new!r}")

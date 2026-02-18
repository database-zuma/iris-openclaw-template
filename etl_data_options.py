#!/usr/bin/env python3
"""
ETL: Parse data-options-cleaned.xlsx → portal.store_display_options
Long/normalized format: area, store_name, display_type, gender, kategori, capacity
"""

import openpyxl
import csv
import io
import sys

EXCEL_PATH = '/Users/database-zuma/.openclaw/workspace/data-options-cleaned.xlsx'
OUTPUT_CSV = '/Users/database-zuma/.openclaw/workspace/data-options-upload.csv'

SKIP_SHEETS = ['Rules']
SKIP_ROW_LABELS = {'Nama Toko', 'Grand Total', None, 'Nama Toko '}

def clean_str(v):
    if v is None:
        return None
    s = str(v).strip()
    return s if s else None

def clean_num(v):
    if v is None:
        return None
    try:
        f = float(v)
        if f == int(f):
            return int(f)
        return round(f, 4)
    except (ValueError, TypeError):
        return None

def parse_regular_sheet(ws, area):
    """
    Regular sheets (Bali 1, Bali 2, Bali 3, Lombok, Jakarta, Jatim, Sumatera, Sulawesi):
    Columns: [label, gender1, kategori1, val1, gender2, kategori2, val2, ...]
    Stores defined in header row where col A = 'Nama Toko'
    """
    rows = list(ws.iter_rows(values_only=True))
    
    # Find the store name row (col A == 'Nama Toko')
    store_row_idx = None
    for i, row in enumerate(rows):
        if clean_str(row[0]) == 'Nama Toko':
            store_row_idx = i
            break
    
    if store_row_idx is None:
        print(f"  WARNING: No 'Nama Toko' row found in {area}", file=sys.stderr)
        return []
    
    store_row = rows[store_row_idx]
    
    # Parse store names from header row
    # Structure: col 0 = 'Nama Toko', then groups of 3: [store_name/None, None/Gender, None/Kategori]
    # Actually: col 1 = store1 name, col 2 = None, col 3 = None, col 4 = store2 name, ...
    stores = {}  # col_group_idx -> store_name (col_group_idx = (col-1)//3)
    for col_idx in range(1, len(store_row)):
        val = clean_str(store_row[col_idx])
        if val and val.upper() not in ('GENDER', 'KATEGORI', 'GRAND TOTAL'):
            group_idx = (col_idx - 1) // 3
            if group_idx not in stores:
                stores[group_idx] = val
    
    print(f"  Stores in {area}: {list(stores.values())}")
    
    # Data starts after the header rows (store_row_idx + 1 or +2)
    # Skip any row with col A = 'Grand Total'
    records = []
    for row in rows[store_row_idx + 1:]:
        label = clean_str(row[0])
        if label is None or label in SKIP_ROW_LABELS or label == 'Grand Total':
            continue
        
        # Parse per-store data
        for group_idx, store_name in stores.items():
            base_col = 1 + group_idx * 3
            if base_col + 2 >= len(row):
                continue
            
            gender = clean_str(row[base_col])
            kategori = clean_str(row[base_col + 1])
            capacity = clean_num(row[base_col + 2])
            
            # Skip if capacity is None or 0 AND no gender/kategori
            if capacity is None and gender is None and kategori is None:
                continue
            
            records.append({
                'area': area,
                'store_name': store_name,
                'display_type': label,
                'gender': gender,
                'kategori': kategori,
                'capacity': capacity,
            })
    
    return records


def parse_consignment_sheet(ws, area):
    """
    Consignment sheet: 
    Row 1: ['Grand Total', total1, total2, ...]
    Row 2: ['Nama Toko', store1, store2, ...]
    Row 3+: [display_type, cap1, cap2, ...]
    No gender/kategori columns - each column = one store
    """
    rows = list(ws.iter_rows(values_only=True))
    
    # Find store name row
    store_row_idx = None
    for i, row in enumerate(rows):
        if clean_str(row[0]) == 'Nama Toko':
            store_row_idx = i
            break
    
    if store_row_idx is None:
        print(f"  WARNING: No 'Nama Toko' row in {area}", file=sys.stderr)
        return []
    
    store_row = rows[store_row_idx]
    
    # Parse stores: col 1, 2, 3, ... each is a store name
    stores = {}  # col_idx -> store_name
    for col_idx in range(1, len(store_row)):
        val = clean_str(store_row[col_idx])
        if val:
            stores[col_idx] = val
    
    print(f"  Stores in {area}: {list(stores.values())}")
    
    records = []
    for row in rows[store_row_idx + 1:]:
        label = clean_str(row[0])
        if label is None or label in SKIP_ROW_LABELS or label == 'Grand Total':
            continue
        
        for col_idx, store_name in stores.items():
            if col_idx >= len(row):
                continue
            capacity = clean_num(row[col_idx])
            if capacity is None:
                continue
            
            records.append({
                'area': area,
                'store_name': store_name,
                'display_type': label,
                'gender': None,
                'kategori': None,
                'capacity': capacity,
            })
    
    return records


def main():
    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
    all_records = []
    
    for sheet_name in wb.sheetnames:
        if sheet_name in SKIP_SHEETS:
            continue
        
        ws = wb[sheet_name]
        print(f"\nParsing sheet: {sheet_name}")
        
        if sheet_name == 'Consignment':
            records = parse_consignment_sheet(ws, sheet_name)
        else:
            records = parse_regular_sheet(ws, sheet_name)
        
        print(f"  -> {len(records)} records")
        all_records.extend(records)
    
    print(f"\nTotal records: {len(all_records)}")
    
    # Write CSV
    fieldnames = ['area', 'store_name', 'display_type', 'gender', 'kategori', 'capacity']
    with open(OUTPUT_CSV, 'w', newline='', encoding='utf-8') as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(all_records)
    
    print(f"CSV written to: {OUTPUT_CSV}")
    
    # Print summary
    from collections import Counter
    area_counts = Counter(r['area'] for r in all_records)
    print("\nRecords per area:")
    for area, count in sorted(area_counts.items()):
        print(f"  {area}: {count}")


if __name__ == '__main__':
    main()
